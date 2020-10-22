# -*- coding:utf-8 _*-

#   author : YOYO
#   time :  2020/10/21 9:59
#   email : youyou.xu@enesource.com
#   project_name :  get_city_jingweidu
#   file_name :  do_excel
#   function：获取excel里的企业名称，以传给查询坐标的接口；将查询的坐标写回excel

from openpyxl import load_workbook


class Datas:
    def __init__(self):
        self.company_id = None
        self.company = None
        self.longitude = None  # 经度
        self.latitude = None      # 纬度


class DoExcel:
    def __init__(self):
        self.name = "E:\桌面\PycharmProjects\get_city_jingweidu\doexcel\company.xlsx"
        self.sheetname = "Sheet1"

    def open_excel(self):
        excel = load_workbook(self.name)
        sheet = excel[self.sheetname]
        return excel, sheet

    def read(self):
        sheet = self.open_excel()[1]
        datas = []
        for i in range(2, sheet.max_row + 1):
            data = Datas()
            data.company_id = sheet.cell(i, 1).value
            data.company = sheet.cell(i, 2).value
            datas.append(data)
        return datas

    def write(self, row, longitude=None, latitude=None, district=None, adcode=None):
        excel, sheet = self.open_excel()
        sheet.cell(row, 3).value = longitude
        sheet.cell(row, 4).value = latitude
        sheet.cell(row, 5).value = district
        sheet.cell(row, 6).value = adcode
        excel.save(self.name)

